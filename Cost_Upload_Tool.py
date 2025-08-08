import os
import pandas as pd
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook
import datetime
import threading
import shutil

def set_progress(value):
    progress["value"] = value
    root.update_idletasks()

def build_site_info_dict(site_file_path):
    try:
        xls = pd.ExcelFile(site_file_path)
        sheet_name = xls.sheet_names[0]  # Use the only sheet
        df = pd.read_excel(xls, sheet_name=sheet_name, dtype=str)

        df.columns = df.columns.str.strip()
        df = df.dropna(subset=["SiteCode"])
        df["SiteCode"] = df["SiteCode"].astype(str).str.zfill(4).str.strip()

        site_info_dict = {
            row["SiteCode"]: {
                "Supplier": str(row.get("Supplier", "")).strip(),
                "ODM": str(row.get("ODM", "")).strip(),
                "MS4 Vendor Code": str(row.get("MS4 Vendor Code", "")).strip(),
            }
            for _, row in df.iterrows()
        }

        return site_info_dict

    except Exception as e:
        messagebox.showerror("Error", f"Error building site info dictionary: {e}")
        return {}

def find_price(part_number, site_code, requested_date, site_info_dict, root_folder, debug_callback=None):
    site_code = str(site_code).zfill(4).strip()
    site_info = site_info_dict.get(site_code)
    forecast_price = "1.50"  # Default forecast price

    if not isinstance(site_info, dict):
        return forecast_price, "", "", "NB-F", ""

    supplier = str(site_info.get("Supplier", "")).strip()
    odm = str(site_info.get("ODM", "")).strip()
    ms4_code = str(site_info.get("MS4 Vendor Code", "")).strip()

    if odm and not supplier and not ms4_code:
        supplier = "TBD"
        ms4_code = "MHP"
    elif not supplier or not odm:
        return forecast_price, supplier, odm, "NB-F", ""

    if pd.isna(requested_date):
        return forecast_price, supplier, odm, "NB-F", ""

    try:
        if not isinstance(requested_date, pd.Timestamp):
            requested_date = pd.to_datetime(requested_date)
        target_month_full = requested_date.strftime("%B").lower()    # e.g., "june"
        target_month_abbr = requested_date.strftime("%b").lower()    # e.g., "jun"
    except Exception:
        return forecast_price, supplier, odm, "NB-F", ""

    # Get the supplier/ODM base path
    base_path = os.path.join(root_folder, supplier, odm)
    # Normalize the path to use consistent separators
    base_path = os.path.normpath(base_path)
    
    if not os.path.exists(base_path):
        # Try alternative path constructions in case of path separator issues
        alt_base_path = root_folder.replace('/', '\\') + '\\' + supplier + '\\' + odm
        alt_base_path = os.path.normpath(alt_base_path)
        if os.path.exists(alt_base_path):
            base_path = alt_base_path
        else:
            return forecast_price, supplier, odm, "NB-F", ""

    # Find all date folders and sort them by date (newest first)
    date_folders = []
    try:
        folder_contents = os.listdir(base_path)
    except Exception as e:
        return forecast_price, supplier, odm, "NB-F", ""
    
    for folder_name in folder_contents:
        folder_path = os.path.join(base_path, folder_name)
        if os.path.isdir(folder_path):
            # Try to parse date from folder name (e.g., "JAN'25", "FEB'25")
            try:
                # Handle formats like "JAN'25" or "JANUARY'25"
                if "'" in folder_name:
                    month_part, year_part = folder_name.split("'")
                    month_part = month_part.strip().upper()
                    year_part = year_part.strip()
                    
                    # Convert month name to number
                    month_map = {
                        'JAN': 1, 'JANUARY': 1,
                        'FEB': 2, 'FEBRUARY': 2,
                        'MAR': 3, 'MARCH': 3,
                        'APR': 4, 'APRIL': 4,
                        'MAY': 5,
                        'JUN': 6, 'JUNE': 6,
                        'JUL': 7, 'JULY': 7,
                        'AUG': 8, 'AUGUST': 8,
                        'SEP': 9, 'SEPTEMBER': 9,
                        'OCT': 10, 'OCTOBER': 10,
                        'NOV': 11, 'NOVEMBER': 11,
                        'DEC': 12, 'DECEMBER': 12
                    }
                    
                    if month_part in month_map:
                        month_num = month_map[month_part]
                        year_num = int(year_part) + (2000 if int(year_part) < 50 else 1900)
                        folder_date = pd.Timestamp(year=year_num, month=month_num, day=1)
                        date_folders.append((folder_date, folder_name, folder_path))
            except:
                # If we can't parse the date, still include the folder but with a very old date
                old_date = pd.Timestamp(year=1900, month=1, day=1)
                date_folders.append((old_date, folder_name, folder_path))
    
    # Sort by date (newest first)
    date_folders.sort(key=lambda x: x[0], reverse=True)
    
    if not date_folders:
        return forecast_price, supplier, odm, "NB-F", ""

    # Rank files: Final > New > others (Initial, etc.)
    def file_score(filename):
        name = filename.lower()
        if "final" in name:
            return 3
        elif "new" in name:
            return 2
        elif "initial" in name:
            return 1
        else:
            return 0

    found_prices = set()
    source_date_folder = None

    # Search through date folders (newest first) until we find the part
    for folder_date, folder_name, folder_path in date_folders:
        all_files = [
            os.path.join(root_dir, file)
            for root_dir, _, files in os.walk(folder_path)
            for file in files
            if file.lower().endswith((".xlsx", ".xls"))
        ]

        if not all_files:
            continue  # No Excel files in this folder, try next
        
        # Sort files by descending priority score
        all_files_sorted = sorted(all_files, key=lambda f: file_score(os.path.basename(f)), reverse=True)

        for file_path in all_files_sorted:
            try:
                df = pd.read_excel(file_path, dtype=str, engine="openpyxl" if file_path.endswith("xlsx") else None)
                df.columns = df.columns.str.strip().str.lower()

                # More flexible part column detection - prioritize specific part columns over generic "item"
                part_col = None
                item_col = None
                
                for col in df.columns:
                    col_lower = col.lower()
                    if (col_lower == "p/n" or 
                        "part number" in col_lower or 
                        "hp part" in col_lower or 
                        "part #" in col_lower or
                        any(keyword in col_lower for keyword in ["part", "sku", "material", "component"])):
                        part_col = col
                        break  # Found a specific part column, use it
                    elif "item" in col_lower and item_col is None:
                        item_col = col  # Keep track of item column as fallback
                
                # If no specific part column found, use item column as fallback
                if part_col is None and item_col is not None:
                    part_col = item_col

                def score_column(col):
                    col_lower = col.lower()
                    score = 0
                    if any(word in col_lower for word in ["price", "cost", "pricing", "rate", "amount", "value", "unit cost", "orderable"]):
                        score += 1
                        if target_month_full in col_lower or target_month_abbr in col_lower:
                            score += 1
                        # Give "orderable" columns very high priority (10 points vs 1-2 for others)
                        if "orderable" in col_lower:
                            score += 10
                    return score

                def extract_date_from_column(col):
                    """Extract date from column name for sorting by recency"""
                    import re
                    import calendar
                    
                    # Look for various date patterns:
                    
                    patterns = [
                        r'([A-Za-z]+)\s+(\d{4})',           # "JULY 2025"
                        r"([A-Za-z]+)'(\d{4})\.",           # "July'2025." (with period after year)
                        r"([A-Za-z]+)'(\d{4})",             # "july'2022" (with apostrophe)
                        r"([A-Za-z]+)\s+'(\d{4})",          # "july '2023" (space before apostrophe)
                        r"([A-Za-z]+)\.(\d{4})\s*\.",       # "may.2025 ." (period after month, space and period after year)
                        r"([A-Za-z]+)\.(\d{4})",            # "July.2025" (with period after month)
                        r'([A-Za-z]+)(\d{4})',              # "jul2025" (no space)
                        r'([A-Za-z]+)\s+(\d{4})\.',         # "July 2025."
                        r"([A-Za-z]+)\s+'(\d{4})\.",        # "July '2025." (with period after year)
                    ]
                    
                    for pattern in patterns:
                        match = re.search(pattern, col)
                        if match:
                            month_name, year_str = match.groups()
                            try:
                                # Try to convert month name to number
                                month_num = None
                                
                                # Try full month names first
                                for i, name in enumerate(calendar.month_name[1:], 1):
                                    if name.upper() == month_name.upper():
                                        month_num = i
                                        break
                                
                                # Try abbreviated month names
                                if month_num is None:
                                    for i, name in enumerate(calendar.month_abbr[1:], 1):
                                        if name.upper() == month_name.upper():
                                            month_num = i
                                            break
                                
                                # Try common variations
                                if month_num is None:
                                    month_variations = {
                                        'JUL': 7, 'AUG': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DEC': 12,
                                        'NOBEMBER': 11,  # Handle typo "nobember"
                                    }
                                    month_num = month_variations.get(month_name.upper())
                                
                                if month_num:
                                    year = int(year_str)
                                    return pd.Timestamp(year=year, month=month_num, day=1)
                            except:
                                continue  # Try next pattern if this one fails
                    
                    return None

                # Get all price columns first
                all_price_columns = [col for col in df.columns if any(w in col.lower() for w in ["price", "cost", "pricing", "rate", "amount", "value", "unit cost", "orderable"])]
                
                # Separate columns with dates from those without
                dated_price_columns = []
                undated_price_columns = []
                
                for col in all_price_columns:
                    col_date = extract_date_from_column(col)
                    if col_date:
                        # Include both column, date, and score for sorting
                        col_score = score_column(col)
                        dated_price_columns.append((col, col_date, col_score))
                    else:
                        undated_price_columns.append(col)
                
                # Sort dated columns first by date (most recent first), then by score (highest first)
                dated_price_columns.sort(key=lambda x: (x[1], x[2]), reverse=True)
                
                # Sort undated columns by regular score
                undated_price_columns_scored = sorted(undated_price_columns, key=score_column, reverse=True)
                
                # Combine: dated columns first (most recent first, then by score), then undated columns
                price_columns = [col for col, _, _ in dated_price_columns] + undated_price_columns_scored

                if part_col and price_columns:
                    # More flexible part number matching
                    part_number_clean = str(part_number).strip().upper().replace('-', '').replace('_', '').replace(' ', '')
                    
                    # Debug: Show cleaned versions of sample parts
                    sample_cleaned = [str(p).strip().upper().replace('-', '').replace('_', '').replace(' ', '') for p in df[part_col].head(10).astype(str)]
                    
                    # Debug: Check if our target part exists anywhere in the file
                    all_parts_cleaned = df[part_col].astype(str).str.strip().str.upper().str.replace('-', '').str.replace('_', '').str.replace(' ', '')
                    target_exists = (all_parts_cleaned == part_number_clean).any()
                    
                    if target_exists:
                        target_row_indices = df[all_parts_cleaned == part_number_clean].index.tolist()
                        for idx in target_row_indices[:3]:  # Show first 3 matches
                            pass
                    
                    matched_rows = df[df[part_col].astype(str).str.strip().str.upper().str.replace('-', '').str.replace('_', '').str.replace(' ', '') == part_number_clean]
                    
                    # If no exact match, try partial matching
                    if matched_rows.empty:
                        matched_rows = df[df[part_col].astype(str).str.strip().str.upper().str.contains(str(part_number).strip().upper(), na=False)]
                    
                    for _, row in matched_rows.iterrows():
                        for price_col in price_columns:
                            price = row.get(price_col)
                            # More flexible price validation
                            if pd.notna(price) and str(price).strip() and str(price).strip().lower() not in ['', 'nan', 'none', 'null', 'n/a', '-']:
                                price_str = str(price).strip()
                                # Remove common non-numeric characters but keep decimal points
                                import re
                                clean_price = re.sub(r'[^\d\.,]', '', price_str)
                                if clean_price and clean_price.replace('.', '').replace(',', '').isdigit():
                                    found_prices.add(price_str)
                                    source_date_folder = folder_name
                                    break  # Use first good column per row

            except Exception:
                continue
        
        # If we found prices in this date folder, stop searching (we want the newest)
        if found_prices:
            break

    if not found_prices:
        return forecast_price, supplier, odm, "NB-F", ""
    elif len(found_prices) == 1:
        return next(iter(found_prices)), supplier, odm, "All", source_date_folder or ""
    else:
        clean_prices = [p for p in found_prices if p is not None]
        return ", ".join(sorted(clean_prices, key=str)), supplier, odm, "All", source_date_folder or ""


def submit():
    media_file = media_file_var.get()
    media_sheet = media_sheet_var.get()
    site_file = site_file_var.get()
    root_folder = root_folder_var.get()
    original_template_path = filedialog.askopenfilename(
        title="Select Cost Upload Template",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )

    if not original_template_path:
        return  # User cancelled

    # Create a timestamped copy
    timestamp = datetime.datetime.now().strftime("%Y%m%d")
    template_dir = os.path.dirname(original_template_path)
    template_name = f"PSO CCS MS4 Cost Upload_{timestamp}.xlsx"
    template_path = os.path.join(template_dir, template_name)

    try:
        shutil.copy(original_template_path, template_path)
    except Exception as e:
        messagebox.showerror("Copy Error", f"Could not create a working copy of the template:\n{e}")
        return

    if not all([media_file, media_sheet, site_file, root_folder, template_path]):
        messagebox.showerror("Input Error", "Please fill out all fields and select the template.")
        return

    try:
        media_df = pd.read_excel(media_file, sheet_name=media_sheet)
        media_df.columns = media_df.columns.str.strip()
        # At the start
        update_progress(5, "Reading Media Tracker...")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to read Media Tracker file: {e}")
        return

    required_cols = ['PartNumber', 'SiteCode', 'Requested Date']
    for col in required_cols:
        if col not in media_df.columns:
            messagebox.showerror("Missing Column", f"Column missing in Media Tracker: {col}")
            return
    # After reading media_df
    update_progress(15, "Validating columns...")

    site_info_dict = build_site_info_dict(site_file)
    # Normalize and filter media_df so only rows with valid site codes remain
    media_df["SiteCode"] = media_df["SiteCode"].astype(str).str.zfill(4).str.strip()
    valid_sitecodes = set(site_info_dict.keys())
    media_df = media_df[media_df["SiteCode"].isin(valid_sitecodes)].reset_index(drop=True)
    media_df = media_df[
        media_df['Comments(Procurement)'].isna() |
        (media_df['Comments(Procurement)'].astype(str).str.strip().replace('nan', '') == '')
    ].reset_index(drop=True)
    # After building site_info_dict and filtering media_df
    update_progress(30, "Looking up prices...")

    def find_price_with_debug(row):
        return find_price(row['PartNumber'], row['SiteCode'], row['Requested Date'], site_info_dict, root_folder)

    results = media_df.apply(find_price_with_debug, axis=1, result_type='expand')
    # After applying find_price
    update_progress(50, "Preparing data for template...")

    media_df['Price'] = results[0]
    media_df['Supplier'] = results[1]
    media_df['ODM'] = results[2]
    media_df['Cost Type'] = results[3]
    media_df['Source Date Folder'] = results[4]
    media_df['MS4 Vendor Code'] = media_df['SiteCode'].map(
        lambda code: site_info_dict.get(str(code).zfill(4), {}).get("MS4 Vendor Code", "")
    )


    # Only update Comments(Procurement) for rows that were searched and recorded in the cost upload template
    try:
        wb_ckit = load_workbook(media_file)
        ws_ckit = wb_ckit[media_sheet]
        # Find the Comments(Procurement), PartNumber, SiteCode, Requested Date column indices
        header_row = None
        for row in ws_ckit.iter_rows(min_row=1, max_row=5):
            for cell in row:
                if cell.value and str(cell.value).strip() == 'Comments(Procurement)':
                    header_row = cell.row
                    break
            if header_row:
                break
        if not header_row:
            header_row = 1  # fallback
        col_map = {}
        for cell in ws_ckit[header_row]:
            if cell.value:
                col_map[str(cell.value).strip()] = cell.column
        # Build a set of keys for the rows that were processed (from media_df)
        processed_keys = set()
        for _, row in media_df.iterrows():
            key = (
                str(row.get('PartNumber', '')).strip().upper(),
                str(row.get('SiteCode', '')).strip().zfill(4),
                str(row.get('Requested Date', '')).strip()
            )
            processed_keys.add(key)
        # Now update only those rows in the Excel file
        for row in ws_ckit.iter_rows(min_row=header_row+1, max_row=ws_ckit.max_row):
            part_val = row[col_map.get('PartNumber', 0)-1].value if 'PartNumber' in col_map else None
            site_val = row[col_map.get('SiteCode', 0)-1].value if 'SiteCode' in col_map else None
            date_val = row[col_map.get('Requested Date', 0)-1].value if 'Requested Date' in col_map else None
            key = (
                str(part_val).strip().upper() if part_val is not None else '',
                str(site_val).strip().zfill(4) if site_val is not None else '',
                str(date_val).strip() if date_val is not None else ''
            )
            if key in processed_keys:
                cell = row[col_map.get('Comments(Procurement)', 0)-1] if 'Comments(Procurement)' in col_map else None
                if cell is not None:
                    # Find the matching DataFrame row and use its comment
                    df_match = media_df[(media_df['PartNumber'].astype(str).str.strip().str.upper() == key[0]) &
                                        (media_df['SiteCode'] == key[1]) &
                                        (media_df['Requested Date'].astype(str).str.strip() == key[2])]
                    if not df_match.empty:
                        cell.value = df_match.iloc[0]['Comments(Procurement)']
                    else:
                        cell.value = "Forecast price has been uploaded to CCS"
        wb_ckit.save(media_file)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to update Comments(Procurement) in doc ckit tracker: {e}")

    # Set Comments(Procurement) based on whether a real cost was found
    def comment_for_row(row):
        price = str(row.get('Price', '')).strip()
        cost_type = str(row.get('Cost Type', '')).strip()
        source_folder = str(row.get('Source Date Folder', '')).strip()
        
        # Only set as cost uploaded if price is found and cost type is not NB-F (forecast)
        if price and cost_type != 'NB-F':
            if source_folder:
                return f"Cost from {source_folder} uploaded to CCS"
            else:
                return "Cost is uploaded to CCS"
        else:
            return "Forecast price has been uploaded to CCS"
    media_df['Comments(Procurement)'] = media_df.apply(comment_for_row, axis=1)
    media_df.fillna("", inplace=True)

    try:
        wb = load_workbook(template_path)
        ws = wb["Input"]
        for row in ws.iter_rows(min_row=1, max_row=20):
            if any(cell.value == "PART NO." for cell in row):
                HEADER_ROW = row[0].row
                break  # Adjusted header row

        # Build headers dictionary from the correct row
        headers = {
            str(cell.value).replace('\n', ' ').strip(): cell.column
            for cell in ws[HEADER_ROW]
            if cell.value is not None
        }

        # Validate required headers exist
        required_headers = [
            "PART NO.", "PART DESCRIPTION", "SUPPLIER NAME", "Site", 
            "Cost (must be in USD)", "Vendor Code", 
            "MKT SHARE %", "Cost Type", "Condition Type"
        ]
        missing_headers = [h for h in required_headers if h not in headers]
        if missing_headers:
            messagebox.showerror("Missing Headers", f"Missing columns in template: {', '.join(missing_headers)}")
            return

        # Fill Vendor Code column if it exists
        vendor_code_cols = [cell.column for cell in ws[HEADER_ROW] if cell.value == "Vendor Code"]


        def safe(val):
            if pd.isna(val):
                return ""
            elif isinstance(val, (pd.Timestamp, pd.Timedelta)):
                return str(val)
            elif isinstance(val, (int, float, str)):
                return val
            else:
                return str(val)

        # Load Admin sheet to build site mapping
        if "Admin" not in wb.sheetnames:
            messagebox.showerror("Error", "Admin sheet not found in template.")
            return

        admin_ws = wb["Admin"]
        site_mapping = {}

        for row in admin_ws.iter_rows(min_row=2, values_only=True):  # Assuming row 1 is header
            site_name = str(row[0]).strip() if row[0] else ""
            site_code = str(row[2]).zfill(4).strip() if row[2] else ""
            if site_code:
                site_mapping[site_code] = site_name
    
        # After loading the workbook
        update_progress(65, "Filling in template fields...")
    
        # Now write to Input sheet
        for i, row in media_df.iterrows():
            excel_row = i + HEADER_ROW + 1  # Write below header row

            site_code = str(row['SiteCode']).zfill(4)
            site_display_name = site_mapping.get(site_code, f"{site_code} UNKNOWN")

            ws.cell(row=excel_row, column=headers["Site"]).value = site_display_name
            ws.cell(row=excel_row, column=headers["PART NO."]).value = safe(row.get("PartNumber"))
            ws.cell(row=excel_row, column=headers["PART DESCRIPTION"]).value = safe(row.get("Description"))
            ws.cell(row=excel_row, column=headers["SUPPLIER NAME"]).value = safe(row.get("Supplier"))
            ws.cell(row=excel_row, column=headers["Cost (must be in USD)"]).value = safe(row.get("Price"))
            vendor_code_value = safe(row.get("MS4 Vendor Code"))
            for col in vendor_code_cols:
                ws.cell(row=excel_row, column=col).value = vendor_code_value
            ws.cell(row=excel_row, column=headers["MKT SHARE %"]).value = "100"
            ws.cell(row=excel_row, column=headers["Cost Type"]).value = safe(row.get("Cost Type"))
            ws.cell(row=excel_row, column=headers["Condition Type"]).value = "PB00"
            ws.cell(row=excel_row, column=headers["EFFECTIVE DATE"]).value = datetime.datetime.today().strftime("%m/%d/%Y")
            if "Comments(Procurement)" in headers:
                ws.cell(row=excel_row, column=headers["Comments(Procurement)"]).value = safe(row.get("Comments(Procurement)"))
            if "Source Date Folder" in headers:
                ws.cell(row=excel_row, column=headers["Source Date Folder"]).value = safe(row.get("Source Date Folder"))

        # After writing all values to Excel
        update_progress(85, "Saving Excel file...")

        wb.save(template_path)
        # After saving successfully
        update_progress(100, "✅ Completed successfully!")
        messagebox.showinfo("Success", f"✅ Cost data written into:\n{template_path}")

        

    except Exception as e:
        messagebox.showerror("Save Error", f"Failed to write to template: {e}")


# === GUI SETUP ===
root = tk.Tk()
root.title("Price Lookup Tool")
root.geometry("600x400")

frame = ttk.Frame(root, padding=20)
frame.pack(expand=True, fill="both")

media_file_var = tk.StringVar()
media_sheet_var = tk.StringVar()
site_file_var = tk.StringVar()
root_folder_var = tk.StringVar()

def browse_file(var):
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if file_path:
        var.set(file_path)

def browse_folder(var):
    folder_path = filedialog.askdirectory()
    if folder_path:
        var.set(folder_path)

def browse_media_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if file_path:
        media_file_var.set(file_path)
        try:
            xls = pd.ExcelFile(file_path)
            sheet_combo["values"] = xls.sheet_names
            if xls.sheet_names:
                media_sheet_var.set(xls.sheet_names[0])  # Default to first sheet
        except Exception as e:
            messagebox.showerror("Error", f"Could not read sheet names:\n{e}")

fields = [
    ("Media Tracker File", media_file_var, lambda: browse_media_file()),
    ("Media Tracker Sheet", media_sheet_var, None),
    ("Site Info File", site_file_var, lambda: browse_file(site_file_var)),
    ("Root Folder (Supplier/ODM/Date)", root_folder_var, lambda: browse_folder(root_folder_var)),
]

for i, (label, var, browse_cmd) in enumerate(fields):
    ttk.Label(frame, text=label).grid(row=i, column=0, sticky="w", pady=5)
    if var == media_sheet_var:
        global sheet_combo
        sheet_combo = ttk.Combobox(frame, textvariable=media_sheet_var, width=43, state="readonly")
        sheet_combo.grid(row=i, column=1, padx=5, pady=5)
    else:
        entry = ttk.Entry(frame, textvariable=var, width=45)
        entry.grid(row=i, column=1, padx=5, pady=5)

    if browse_cmd:
        ttk.Button(frame, text="Browse", command=browse_cmd).grid(row=i, column=2, padx=5)


# Progress bar (loading bar)
progress = ttk.Progressbar(frame, orient="horizontal", mode="determinate", length=300, maximum=100)
progress.grid(row=len(fields)+1, column=1, pady=10)
progress_label = ttk.Label(frame, text="")
progress_label.grid(row=len(fields)+2, column=1, pady=5)

def update_progress(stage, message):
    progress["value"] = stage
    progress_label.config(text=message)
    root.update_idletasks()


# Function to run submission with loading bar
def run_submit():
    def task():
        try:
            update_progress(0, "Initializing...")
            submit()
        finally:
            progress["value"] = 0
            progress_label.config(text="")
    threading.Thread(target=task).start()

# Submit button
ttk.Button(frame, text="Submit", command=run_submit).grid(row=len(fields), column=1, pady=20)

root.mainloop()