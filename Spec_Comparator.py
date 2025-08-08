import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import difflib
import re
import calendar
import datetime

def extract_date_from_col(col):
    """
    Extracts a date from a column name string. Supports formats like May'25, May 2025, 2025-05, etc.
    Returns a datetime.date object or None if no date found.
    """
    col = str(col)
    m = re.search(r"([A-Za-z]+)[\s'.]*(\d{2,4})", col)
    if m:
        month_str, year_str = m.group(1), m.group(2)
        try:
            month = list(calendar.month_abbr).index(month_str[:3].title()) if month_str[:3].title() in calendar.month_abbr else list(calendar.month_name).index(month_str.title())
        except:
            month = 1
        year = int(year_str)
        if year < 100:
            year += 2000
        return datetime.date(year, month, 1)
    m2 = re.search(r"(\d{4})[-/](\d{1,2})", col)
    if m2:
        year, month = int(m2.group(1)), int(m2.group(2))
        return datetime.date(year, month, 1)
    return None

def find_spec_columns(columns):
    # First, exclude columns that are clearly identifiers/indexes
    excluded_patterns = ["item", "index", "id", "number"]
    
    # Look for actual spec content columns, prioritizing "Specs" over "Spec Items"
    spec_patterns = [
        r"^specs$",           # Exact match for "Specs"
        r"^specifications$",   # Exact match for "Specifications"
        r"^spec$",            # Exact match for "Spec" (but lower priority)
        r"specs",             # Contains "specs"
        r"specifications",    # Contains "specifications"
        r"spec"               # Contains "spec" (lowest priority)
    ]
    
    found_columns = []
    
    # Search in priority order
    for pattern in spec_patterns:
        for col in columns:
            col_lower = col.strip().lower()
            # Skip if it contains identifier keywords
            if any(excl in col_lower for excl in excluded_patterns):
                continue
            # Check if it matches the current pattern
            if re.search(pattern, col_lower):
                found_columns.append(col)
                break  # Found one for this pattern, move to next
        if found_columns:
            break  # Found a column, stop searching
    
    return found_columns


def normalize_spec_string(s):
    # Convert to string
    s = str(s)
    # Replace HTML and Unicode non-breaking spaces with a regular space
    s = s.replace('\xa0', ' ').replace('&nbsp;', ' ')
    # Replace full-width semicolons/colons with ASCII
    s = s.replace('；', ';').replace('：', ':')
    # Remove zero-width and invisible Unicode characters
    s = re.sub(r'[\u200b-\u200f\u202a-\u202e\u2060-\u206f]', '', s)
    # Replace all whitespace (spaces, tabs, etc.) with a single space
    s = re.sub(r'\s+', ' ', s)
    # Remove spaces around punctuation like ; :
    s = re.sub(r'\s*([;:])\s*', r'\1', s)
    # Strip and uppercase
    return s.strip().upper()

def extract_kv_pairs(s):
    s = str(s)
    s = re.sub(r'[^\w\s\.:=xX±\-]', '', s)
    s = re.sub(r'\s+', ' ', s)

    # Common pattern for dimensions like 5x6x3 or 5*6*3
    dimension_pattern = r'(\w+)\s*[:=]?\s*([\d\.]+[xX\*][\d\.]+(?:[xX\*][\d\.]+)?)'
    dimension_matches = re.findall(dimension_pattern, s)

    kv = {k.lower(): v.replace(' ', '').upper() for k, v in dimension_matches}

    # Additional key-value pattern
    pattern = r'(\w+)\s*[:=]\s*([\w\.xX±\-]+)'
    for k, v in re.findall(pattern, s):
        if k.lower() not in kv:
            kv[k.lower()] = v.replace(' ', '').upper()

    return kv


def extract_quantity_number(qty_str):
    try:
        s = str(qty_str).strip().lower().replace("pcs", "").replace(",", "")
        if 'k' in s:
            return int(float(s.replace('k', '')) * 1000)
        return int(float(s))
    except:
        return None

def is_quantity_column(col_name):
    return extract_quantity_number(col_name) is not None

def find_price_columns(columns, keywords=None):
    if keywords is None:
        keywords = ["price", "cost", "pricing", "unit cost", "unit price", "orderable price"]
    return [col for col in columns if any(k in col.lower() for k in keywords)]

def match_specs_and_append_prices(quote_df, specs_folder):
    # Dynamically detect all spec columns in the quote file
    spec_col_candidates = [col for col in quote_df.columns if "spec" in col.strip().lower()]
    if not spec_col_candidates:
        raise KeyError("No spec column found in quote_df. Expected a column containing 'spec'.")
    # We'll search all spec columns for each row

    # Check for Spec Line column
    spec_line_col = None
    for col in quote_df.columns:
        if col.strip().lower() == "spec line":
            spec_line_col = col
            break

    all_spec_files = [f for f in os.listdir(specs_folder) if f.endswith((".xls", ".xlsx", ".xlsb"))]
    result_df = quote_df.copy()
    added_columns = []

    for spec_file in all_spec_files:
        file_path = os.path.join(specs_folder, spec_file)
        try:
            if spec_file.endswith(".xlsb"):
                xls = pd.ExcelFile(file_path, engine="pyxlsb")
            else:
                xls = pd.ExcelFile(file_path)
        except Exception:
            continue

        for sheet_name in xls.sheet_names:
            try:
                df = xls.parse(sheet_name)
                df.columns = [str(c).strip() for c in df.columns]
            except Exception:
                continue

            # Accept any column containing "spec"
            spec_cols = find_spec_columns(df.columns)
            if not spec_cols:
                continue

            price_cols = find_price_columns(df.columns)

            matched_prices = []
            matched_volumes = []
            matched_mask = []

            for _, row in quote_df.iterrows():
                # For each quote row, search all quote spec columns for the best match in all spec columns of the spec file
                price_found = None
                vol_found = None
                matched = False

                # Detect if this is a "volume table" style spec file
                has_volume_col = any("volume" in col.lower() for col in df.columns)
                qty_cols = [col for col in df.columns if re.match(r"^\d+(\.\d+)?k$", col.strip().lower())]

                if has_volume_col or qty_cols:
                    # --- "Volume Table" style: require exact match on spec and volume ---
                    for quote_spec_col in spec_col_candidates:
                        quote_spec = normalize_spec_string(row[quote_spec_col])
                        for spec_col_in_file in spec_cols:
                            norm_specs = df[spec_col_in_file].astype(str).apply(normalize_spec_string)
                            matches = df[norm_specs == quote_spec]
                            if not matches.empty:
                                # Try to match volume
                                quote_volume = None
                                # Try to get the quote's volume from the quote row (from any volume column)
                                for vcol in quote_df.columns:
                                    if "volume" in vcol.lower() and pd.notna(row[vcol]):
                                        quote_volume = extract_quantity_number(row[vcol])
                                        if quote_volume is not None:
                                            break
                                # Try to match volume columns with dates
                                if has_volume_col and quote_volume is not None:
                                    # Find volume columns (containing "volume")
                                    volume_cols = [col for col in df.columns if "volume" in col.lower()]
                                    for vol_col in volume_cols:
                                        vol_matches = matches[matches[vol_col].apply(lambda x: extract_quantity_number(x) == quote_volume)]
                                        if not vol_matches.empty:
                                            # Find corresponding pricing column
                                            # Look for a pricing column with similar name pattern
                                            pricing_col = vol_col.replace("Volume", "pricing").replace("volume", "pricing")
                                            if pricing_col in df.columns:
                                                price_found = vol_matches.iloc[0][pricing_col]
                                                vol_found = vol_matches.iloc[0][vol_col]
                                                matched = True
                                                break
                                            else:
                                                # Try to find any pricing column
                                                price_cols_in_file = find_price_columns(df.columns)
                                                if price_cols_in_file:
                                                    price_found = vol_matches.iloc[0][price_cols_in_file[0]]
                                                    vol_found = vol_matches.iloc[0][vol_col]
                                                    matched = True
                                                    break
                                    if matched:
                                        break
                                # Try to match quantity columns (e.g., "1K", "5K")
                                elif qty_cols and quote_volume is not None:
                                    # Find the closest quantity column
                                    qty_numbers = [(col, extract_quantity_number(col)) for col in qty_cols]
                                    qty_numbers = [(col, num) for col, num in qty_numbers if num is not None]
                                    if qty_numbers:
                                        closest_col, _ = min(qty_numbers, key=lambda x: abs(x[1] - quote_volume))
                                        try:
                                            price_found = matches.iloc[0][closest_col]
                                            vol_found = closest_col
                                            matched = True
                                            break
                                        except:
                                            continue
                else:
                    # --- Fuzzy matching as before ---
                    best_score = 0
                    best_row = None
                    best_col = None
                    best_quote_spec_val = None

                    for quote_spec_col in spec_col_candidates:
                        quote_spec = normalize_spec_string(row[quote_spec_col])
                        for spec_col_in_file in spec_cols:
                            norm_specs = df[spec_col_in_file].astype(str).apply(normalize_spec_string)
                            for idx, spec_val in norm_specs.items():
                                score = difflib.SequenceMatcher(None, quote_spec, spec_val).ratio()
                                if score > best_score:
                                    best_score = score
                                    best_row = df.loc[idx]
                                    best_col = spec_col_in_file
                                    best_quote_spec_val = quote_spec

                    if best_score > 0.85 and best_row is not None:
                        for pcol in price_cols:
                            try:
                                price_found = float(best_row[pcol])
                                break
                            except:
                                continue
                        for vcol in best_row.index:
                            if any(x in vcol.lower() for x in ["moq", "volume", "qty", "quantity"]):
                                vol_found = best_row[vcol]
                                break
                        matched = True

                matched_prices.append(price_found)
                matched_volumes.append(vol_found)
                matched_mask.append(matched)
            if any(pd.notna(p) for p in matched_prices):
                # Use file name without extension and sheet name for column naming
                file_name_base = os.path.splitext(spec_file)[0]
                matched_price_col = f"{file_name_base} - {sheet_name} Matched Price"
                result_df[matched_price_col] = matched_prices
                # Insert the volume column right after the matched price column
                vol_col_name = f"{file_name_base} - {sheet_name} Volume"
                col_list = list(result_df.columns)
                price_idx = col_list.index(matched_price_col)
                result_df.insert(price_idx + 1, vol_col_name, matched_volumes)
                # Insert Cost Delta column after volume column
                cost_delta_col_name = f"{file_name_base} - {sheet_name} Cost Delta"
                # Find all quote price columns with month/year info
                def norm_col_name(col):
                    return re.sub(r'[^a-zA-Z0-9]', '', str(col)).lower()
                price_cols_quote = [col for col in quote_df.columns if any(x in norm_col_name(col) for x in ["price", "cost", "pricing"])]
                dated_cols = [(col, extract_date_from_col(col)) for col in price_cols_quote]
                dated_cols = [(col, dt) for col, dt in dated_cols if dt is not None]
                if dated_cols:
                    quote_price_col = max(dated_cols, key=lambda x: x[1])[0]
                elif price_cols_quote:
                    quote_price_col = price_cols_quote[0]
                else:
                    quote_price_col = None

                cost_deltas = []
                for i, price in enumerate(matched_prices):
                    quote_price = None
                    if quote_price_col:
                        try:
                            quote_price = float(quote_df.iloc[i][quote_price_col])
                        except:
                            quote_price = None
                    sheet_price = None
                    try:
                        sheet_price = float(matched_prices[i])
                    except:
                        sheet_price = None
                    if sheet_price is not None and quote_price is not None:
                        cost_deltas.append(round(quote_price - sheet_price, 4))
                    else:
                        cost_deltas.append(None)
                result_df.insert(price_idx + 2, cost_delta_col_name, cost_deltas)
                added_columns.append(matched_price_col)
                added_columns.append(vol_col_name)
                added_columns.append(cost_delta_col_name)

    # Move 'Remark' and any spec columns to end
    for col in ["Remark"] + spec_col_candidates:
        if col in result_df.columns:
            result_df = result_df[[c for c in result_df.columns if c != col] + [col]]

    return result_df, added_columns

def bold_columns(ws, header, original_cols, extra_bold_cols=None):
    from openpyxl.styles import Font
    bold_cols = [idx for idx, col in enumerate(header) if any(col.lower() == orig_col.lower() for orig_col in original_cols)]
    if extra_bold_cols:
        bold_cols += extra_bold_cols
    bold_cols = list(set(bold_cols))
    bold_font = Font(bold=True)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for idx in bold_cols:
            if idx < len(row):
                row[idx].font = bold_font
    for idx in bold_cols:
        ws.cell(row=1, column=idx+1).font = bold_font

def highlight_prices(file_path, sheet_name, extra_quote_price_keywords=None):

    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # Lowest
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # Highest
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Tie
    blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")    # Medium

    header = [str(cell.value).strip() for cell in ws[1]]

    price_col_indices = set()
    if extra_quote_price_keywords:
        for i, h in enumerate(header):
            if any(k.lower() in h.lower() for k in extra_quote_price_keywords):
                price_col_indices.add(i + 1)

    # Include sheet price columns (skip identifiers and volume columns)
    skip_keywords = ["Item", "HPPart#", "Type", "Remark", "SPECs"]
    for i, h in enumerate(header):
        if (
            h not in skip_keywords
            and "volume" not in h.lower()
            and "variance" not in h.lower()
            and "cost delta" not in h.lower()
            and "confidence" not in h.lower()
        ):
            price_col_indices.add(i + 1)

    price_col_indices = sorted(price_col_indices)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        values = []
        for idx in price_col_indices:
            cell = row[idx - 1]
            try:
                val = round(float(str(cell.value).strip()), 4)
                if val == 0:
                    continue  # Skip zero values for highlighting
                values.append((val, cell))
            except:
                continue

        if not values:
            continue

        vals_only = [val for val, _ in values]
        unique_vals = sorted(set(vals_only))

        if len(values) == 1:
            values[0][1].fill = yellow_fill
        elif len(set(vals_only)) == 1:
            for _, cell in values:
                cell.fill = yellow_fill
        else:
            min_val = min(vals_only)
            max_val = max(vals_only)
            for val, cell in values:
                if val == min_val:
                    cell.fill = green_fill
                elif val == max_val:
                    cell.fill = red_fill
                elif vals_only.count(val) > 1:
                    cell.fill = blue_fill
                else:
                    cell.fill = yellow_fill

    wb.save(file_path)

def kv_score(kv1, kv2):
    if not kv1 or not kv2:
        return 0

    kv1 = {k.strip().lower(): v.strip().lower() for k, v in kv1.items()}
    kv2 = {k.strip().lower(): v.strip().lower() for k, v in kv2.items()}

    keys = set(kv1.keys()).union(set(kv2.keys()))
    total_keys = len(keys)
    if total_keys == 0:
        return 0

    matches = 0
    for k in keys:
        v1 = kv1.get(k)
        v2 = kv2.get(k)
        if v1 and v2:
            if v1 == v2:
                matches += 1
            else:
                # Soft penalty for mismatch
                matches += 0
        else:
            # Key only in one of them
            matches += 0

    return matches / total_keys


def find_closest_spec_and_costs(quote_spec, specs_folder):
    best_match = None
    best_score = 0
    best_row = None
    best_file = None
    best_sheet = None
    best_part_number = None

    all_spec_files = [f for f in os.listdir(specs_folder) if f.endswith((".xls", ".xlsx", ".xlsb"))]

    norm_quote_spec = normalize_spec_string(quote_spec)
    for spec_file in all_spec_files:
        file_path = os.path.join(specs_folder, spec_file)
        try:
            xls = pd.ExcelFile(file_path)
        except Exception:
            continue

        for sheet_name in xls.sheet_names:
            try:
                df = xls.parse(sheet_name)
                df.columns = [str(c).strip() for c in df.columns]
            except Exception:
                continue

            spec_cols = find_spec_columns(df.columns)
            if not spec_cols:
                continue
            spec_col = spec_cols[0]

            part_number_cols = [
                col for col in df.columns
                if any(
                    kw in col.strip().lower()
                    for kw in ["part number", "partnumber", "p/n", "part", "hppart#"]
                )
            ]
            part_number_col = part_number_cols[0] if part_number_cols else None

            quote_kv = extract_kv_pairs(norm_quote_spec)
            for idx, spec_val in df[spec_col].dropna().items():
                spec_str = normalize_spec_string(spec_val)
                spec_kv = extract_kv_pairs(spec_str)

                # Calculate confidence based on original strings for all matches
                base_score = difflib.SequenceMatcher(None, quote_spec, str(spec_val)).ratio()
                
                # --- Check for exact normalized match ---
                if spec_str == norm_quote_spec:
                    # For exact normalized matches, use original string similarity as confidence
                    score = base_score
                else:
                    # For fuzzy matches, combine original string similarity with key-value matching
                    kv_matches = 0
                    kv_total = max(len(quote_kv), 1)
                    for k, v in quote_kv.items():
                        if k in spec_kv:
                            try:
                                v1 = float(re.sub(r'[^\d\.]', '', v))
                                v2 = float(re.sub(r'[^\d\.]', '', spec_kv[k]))
                                if abs(v1 - v2) < 0.1:
                                    kv_matches += 1
                                    continue
                            except:
                                pass
                            if v == spec_kv[k]:
                                kv_matches += 1
                    kv_sim = kv_score(quote_kv, spec_kv)
                    # Combine base similarity of original strings with key-value pair matching
                    score = min(base_score * 0.7 + kv_sim * 0.3, 1.0)
                    if score < 0.5 and kv_sim > 0.5:
                        score = 0.5 + kv_sim * 0.5

                # Keep track of the best match found so far
                if score > best_score:
                    best_score = score
                    best_match = str(spec_val)  # Store original format, not normalized
                    best_row = df.loc[idx]
                    best_file = spec_file
                    best_sheet = sheet_name
                    best_part_number = str(df.loc[idx][part_number_col]) if part_number_col else None

    return best_match, best_file, best_sheet, best_part_number, best_score

def extract_numbers(s):
    return [float(x) for x in re.findall(r"\d+(?:\.\d+)?", str(s))]

def get_first_price_for_spec(spec, specs_folder):
    all_spec_files = [f for f in os.listdir(specs_folder) if f.endswith((".xls", ".xlsx", ".xlsb"))]
    norm_spec = normalize_spec_string(spec)
    for spec_file in all_spec_files:
        file_path = os.path.join(specs_folder, spec_file)
        try:
            xls = pd.ExcelFile(file_path)
        except Exception:
            continue
        for sheet_name in xls.sheet_names:
            try:
                df = xls.parse(sheet_name)
                df.columns = [str(c).strip() for c in df.columns]
            except Exception:
                continue
            # Accept all forms of 'Spec' columns
            spec_col_patterns = [
                r"^spec$", r"^specs$", r"^specification$", r"^specifications$"
            ]
            spec_cols = find_spec_columns(df.columns)
            if not spec_cols:
                continue
            spec_col = spec_cols[0]
            norm_specs = df[spec_col].astype(str).apply(normalize_spec_string)
            matches = df[norm_specs == norm_spec]
            if not matches.empty:
                # Prioritize columns with 'orderable' in their name for price/cost/pricing
                orderable_cols = [col for col in matches.columns if 'orderable' in col.lower() and (('price' in col.lower()) or ('cost' in col.lower()) or ('pricing' in col.lower()))]
                for col in orderable_cols:
                    try:
                        price = float(matches.iloc[0][col])
                        # Use None for volume if not a quantity column
                        return None, price
                    except:
                        continue
                # If no orderable price/cost/pricing found, try all price/cost/pricing columns
                price_cols = find_price_columns(matches.columns)
                for col in price_cols:
                    try:
                        price = float(matches.iloc[0][col])
                        return None, price
                    except:
                        continue
                # If no price/cost/pricing found, try all quantity columns for a valid price
                qty_cols = [col for col in matches.columns if is_quantity_column(col)]
                for col in qty_cols:
                    try:
                        price = float(matches.iloc[0][col])
                        volume = extract_quantity_number(col)
                        return volume, price
                    except:
                        continue
                # If no price found, fall back to closest quantity column
                qty_numbers = [(col, extract_quantity_number(col)) for col in qty_cols]
                qty_numbers = [(col, num) for col, num in qty_numbers if num is not None]
                if qty_numbers:
                    # Try all columns sorted by volume (ascending)
                    for col, vol in sorted(qty_numbers, key=lambda x: x[1]):
                        try:
                            price = float(matches.iloc[0][col])
                            return vol, price
                        except:
                            continue
    return None, None

def get_closest_price_for_spec(spec, quote_volume, specs_folder):
    """
    Returns (closest_qty_col, price) for the closest quantity column to quote_volume for the given spec.
    If a 'price' or 'cost' column exists, returns its value for the matching spec row.
    """
    all_spec_files = [f for f in os.listdir(specs_folder) if f.endswith((".xls", ".xlsx", ".xlsb"))]
    norm_spec = normalize_spec_string(spec)
    for spec_file in all_spec_files:
        file_path = os.path.join(specs_folder, spec_file)
        try:
            xls = pd.ExcelFile(file_path)
        except Exception:
            continue
        for sheet_name in xls.sheet_names:
            try:
                df = xls.parse(sheet_name)
                df.columns = [str(c).strip() for c in df.columns]
            except Exception:
                continue
            # Accept all forms of 'Spec' columns
            spec_col_patterns = [
                r"^spec$", r"^specs$", r"^specification$", r"^specifications$"
            ]
            spec_cols = find_spec_columns(df.columns)
            if not spec_cols:
                continue
            spec_col = spec_cols[0]
            norm_specs = df[spec_col].astype(str).apply(normalize_spec_string)
            matches = df[norm_specs == norm_spec]
            if not matches.empty:
                # Accept any column containing "price" or "cost" (case-insensitive)
                price_col_candidates = find_price_columns(matches.columns)
                # If columns have date info, pick the most recent
                # Use top-level extract_date_from_col

                dated_cols = [(col, extract_date_from_col(col)) for col in price_col_candidates]
                dated_cols = [(col, dt) for col, dt in dated_cols if dt is not None]
                if dated_cols:
                    # Pick the most recent date
                    most_recent_col = max(dated_cols, key=lambda x: x[1])[0]
                    try:
                        price = float(matches.iloc[0][most_recent_col])
                        return most_recent_col, price
                    except:
                        pass
                elif price_col_candidates:
                    # Fallback: just use the first price/cost column
                    price_col = price_col_candidates[0]
                    try:
                        price = float(matches.iloc[0][price_col])
                        return price_col, price
                    except:
                        pass
                # Fallback: original logic for quantity columns
                qty_cols = [col for col in matches.columns if is_quantity_column(col)]
                if not qty_cols:
                    continue
                qty_numbers = [(col, extract_quantity_number(col)) for col in qty_cols]
                qty_numbers = [(col, num) for col, num in qty_numbers if num is not None]
                if not qty_numbers:
                    continue
                if quote_volume is not None:
                    closest_col, _ = min(qty_numbers, key=lambda x: abs(x[1] - quote_volume))
                else:
                    closest_col = qty_numbers[0][0]
                try:
                    price = float(matches.iloc[0][closest_col])
                    return closest_col, price
                except:
                    continue
    return None, None

# -------- GUI Implementation --------
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

def run_comparator(quote_path, specs_folder, output_path, progress_label=None):
    if not os.path.isfile(quote_path) or not os.path.isdir(specs_folder):
        if progress_label:
            progress_label.config(text="❌ Invalid quote file or specs folder path.")
        else:
            print("❌ Invalid quote file or specs folder path.")
        return
    if progress_label:
        progress_label.config(text="Reading quote file...")
        progress_label.update_idletasks()
    quote_df = pd.read_excel(quote_path)
    result_df, added_cols = match_specs_and_append_prices(quote_df, specs_folder)
    if progress_label:
        progress_label.config(text="Saving matched parts...")
        progress_label.update_idletasks()
    result_df.to_excel(output_path, index=False)

    # --- Create a new sheet for unmatched parts ---

    # --- Create a new sheet for unmatched parts ---
    if added_cols:
        unmatched_mask = result_df[added_cols].isna().all(axis=1)
        unmatched_df = result_df[unmatched_mask]
        matched_df = result_df[~unmatched_mask]
    else:
        unmatched_df = result_df.copy()
        matched_df = result_df.iloc[0:0]

    # --- For each unmatched part, find closest spec ---
    closest_specs = []
    closest_part_numbers = []
    spec_files = []
    spec_sheets = []
    closest_volumes = []
    existing_prices = []
    closest_spec_moq_vols = []
    cost_deltas = []  # New: store cost delta values
    confidence_scores = []  # New: store confidence scores

    # Dynamically find the spec column name (case-insensitive, matches "spec" or "specs")
    spec_col_candidates = find_spec_columns(unmatched_df.columns)
    if spec_col_candidates:
        spec_col = spec_col_candidates[0]
    else:
        raise KeyError("No spec column found in unmatched_df. Expected one of: 'Spec', 'Specs', 'SPEC', 'SPECs'")

    # Find most recent price column in quote file
    # Use top-level extract_date_from_col
    def norm_col_name(col):
        return re.sub(r'[^a-zA-Z0-9]', '', str(col)).lower()
    price_cols = [col for col in unmatched_df.columns if any(x in norm_col_name(col) for x in ["price", "cost", "pricing"])]
    dated_cols = [(col, extract_date_from_col(col)) for col in price_cols]
    dated_cols = [(col, dt) for col, dt in dated_cols if dt is not None]
    if dated_cols:
        most_recent_quote_price_col = max(dated_cols, key=lambda x: x[1])[0]
    elif price_cols:
        most_recent_quote_price_col = price_cols[0]
    else:
        most_recent_quote_price_col = None

    for _, row in unmatched_df.iterrows():
        quote_spec_original = row[spec_col]  # Keep original format
        best_match, best_file, best_sheet, best_part_number, confidence_score = find_closest_spec_and_costs(quote_spec_original, specs_folder)
        closest_specs.append(best_match)
        closest_part_numbers.append(best_part_number)
        spec_files.append(best_file)
        spec_sheets.append(best_sheet)
        confidence_scores.append(f"{round(confidence_score * 100, 1)}%" if confidence_score is not None else None)  # Convert to percentage with % sign

        # Find quote part's volume (try to get the first numeric value from any column with 'volume' in its name)
        quote_volume = None
        for col in unmatched_df.columns:
            if "volume" in col.lower() and pd.notna(row[col]):
                quote_volume = extract_quantity_number(row[col])
                if quote_volume is not None:
                    break

        # Get closest price and volume column
        closest_vol_col, existing_price = get_closest_price_for_spec(best_match, quote_volume, specs_folder)

        # If the column is a price/cost column, don't treat it as a volume column
        if closest_vol_col and ("price" in closest_vol_col.lower() or "cost" in closest_vol_col.lower()):
            closest_volumes.append("")  # or None
        else:
            closest_volumes.append(closest_vol_col)

        # --- Find the MOQ/Volume value for the closest spec ---
        moq_vol_value = None
        if best_match and best_file and best_sheet:
            file_path = os.path.join(specs_folder, best_file)
            try:
                xls = pd.ExcelFile(file_path)
                df = xls.parse(best_sheet)
                df.columns = [str(c).strip() for c in df.columns]
                spec_cols = find_spec_columns(df.columns)
                if spec_cols:
                    spec_col_in_file = spec_cols[0]
                    norm_specs = df[spec_col_in_file].astype(str).apply(normalize_spec_string)
                    match_row = df[norm_specs == normalize_spec_string(best_match)]
                    if not match_row.empty:
                        # Look for a column with 'moq', 'volume', or 'qty' in the name first
                        moq_vol_cols = [col for col in df.columns if any(x in col.lower() for x in ["moq", "volume", "qty", "quantity"])]
                        for col in moq_vol_cols:
                            val = match_row.iloc[0][col]
                            if pd.notna(val):
                                moq_vol_value = val
                                break
                        # If no dedicated MOQ/Volume column found, use the closest volume column info
                        if moq_vol_value is None and closest_vol_col:
                            # If closest_vol_col is a quantity column like "1K", "5K", extract the number
                            extracted_qty = extract_quantity_number(closest_vol_col)
                            if extracted_qty:
                                moq_vol_value = f"{extracted_qty:,} pcs"  # Format like "1,000 pcs"
                            else:
                                moq_vol_value = closest_vol_col  # Use as-is if not a quantity
            except Exception:
                pass
        closest_spec_moq_vols.append(moq_vol_value)

        existing_prices.append(existing_price)

        # Calculate cost delta between most recent quote price and existing price
        quote_price = None
        if most_recent_quote_price_col:
            try:
                quote_price = float(row[most_recent_quote_price_col])
            except:
                quote_price = None
        existing_price = existing_prices[-1] if existing_prices else None
        if existing_price is not None and quote_price is not None:
            cost_delta = round(quote_price - existing_price, 4)
        else:
            cost_delta = None
        cost_deltas.append(cost_delta)

    # Insert Closest Part Number before Closest Spec
    insert_idx = list(unmatched_df.columns).index("Closest Spec") if "Closest Spec" in unmatched_df.columns else len(unmatched_df.columns)
    unmatched_df.insert(insert_idx, "Closest Part Number", closest_part_numbers)
    unmatched_df.loc[:, "Closest Spec"] = closest_specs
    unmatched_df.loc[:, "Confidence Score (%)"] = confidence_scores  # New column for confidence scores
    
    # Only add "Closest Spec MOQ/Volume" if there's no "Closest Volume" column with values
    if not any(closest_volumes):
        unmatched_df.loc[:, "Closest Spec MOQ/Volume"] = closest_spec_moq_vols  # New column for MOQ/Volume
    
    unmatched_df.loc[:, "Existing Price"] = existing_prices
    # Insert Cost Delta column right after Existing Price
    if "Existing Price" in unmatched_df.columns:
        cost_delta_idx = list(unmatched_df.columns).index("Existing Price") + 1
        unmatched_df.insert(cost_delta_idx, "Cost Delta", cost_deltas)
    else:
        unmatched_df["Cost Delta"] = cost_deltas
    # Only add "Closest Volume" if there is at least one non-empty value
    if any(closest_volumes):
        unmatched_df.loc[:, "Closest Volume"] = closest_volumes
    elif "Closest Volume" in unmatched_df.columns:
        unmatched_df = unmatched_df.drop(columns=["Closest Volume"])
    unmatched_df.loc[:, "Spec Source File"] = spec_files
    unmatched_df.loc[:, "Spec Source Sheet"] = spec_sheets

    # Remove matched parts' sheet columns from unmatched_df
    if added_cols:
        unmatched_df = unmatched_df.drop(columns=added_cols, errors="ignore")

    # --- Save only matched parts to the first sheet ---
    matched_df.to_excel(output_path, index=False, sheet_name="Matched Parts")

    # --- Bolden all columns from the quote file in the 'Matched Parts' sheet (after writing matched_df) ---
    wb = load_workbook(output_path)
    if 'Matched Parts' in wb.sheetnames:
        ws_matched = wb['Matched Parts']
        header = [str(cell.value).strip() for cell in ws_matched[1]]
        original_quote_df = pd.read_excel(quote_path)
        original_cols = [c.strip() for c in original_quote_df.columns]
        bold_columns(ws_matched, header, original_cols)
        wb.save(output_path)

    def get_diff_words(a, b):
        a_words = str(a).split()
        b_words = str(b).split()
        a_words_lower = [w.lower() for w in a_words]
        b_words_lower = [w.lower() for w in b_words]
        matcher = difflib.SequenceMatcher(None, a_words_lower, b_words_lower)
        diff_a = []
        diff_b = []
        for tag, i1, i2, j1, j2 in matcher.get_opcodes():
            # Only include words that are truly different (ignoring case)
            if tag == 'replace' or tag == 'insert':
                for idx in range(j1, j2):
                    if b_words_lower[idx] not in a_words_lower[i1:i2]:
                        diff_b.append(b_words[idx])
            if tag == 'replace' or tag == 'delete':
                for idx in range(i1, i2):
                    if a_words_lower[idx] not in b_words_lower[j1:j2]:
                        diff_a.append(a_words[idx])
        return ' '.join(diff_b), ' '.join(diff_a)

    # Add these columns BEFORE writing unmatched_df to Excel
    diff_closest = []
    diff_original = []
    for _, row in unmatched_df.iterrows():
        if pd.notna(row["Closest Spec"]) and pd.notna(row[spec_col]):
            d_closest, d_original = get_diff_words(row[spec_col], row["Closest Spec"])
            diff_closest.append(d_closest)
            diff_original.append(d_original)
        else:
            diff_closest.append("")
            diff_original.append("")

    unmatched_df["Spec Difference (Original Spec)"] = diff_original
    unmatched_df["Spec Difference (Closest Spec)"] = diff_closest

    # --- Create a new sheet for unmatched parts ---
    with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        unmatched_df.to_excel(writer, sheet_name="Unmatched Parts", index=False)

    # --- Bolden all columns from the quote file in the 'Unmatched Parts' sheet (after writing unmatched_df) ---
    wb = load_workbook(output_path)
    if 'Unmatched Parts' in wb.sheetnames:
        ws_unmatched = wb['Unmatched Parts']
        header = [str(cell.value).strip() for cell in ws_unmatched[1]]
        original_quote_df = pd.read_excel(quote_path)
        original_cols = [c.strip() for c in original_quote_df.columns]
        extra_bold_cols = []
        if 'Spec Difference (Original Spec)' in header:
            extra_bold_cols.append(header.index('Spec Difference (Original Spec)'))
        bold_columns(ws_unmatched, header, original_cols, extra_bold_cols)
        wb.save(output_path)

    known_quote_price_keywords = ["pricing", "price", "cost", "costs", "quote", "quoted"]

    highlight_prices(output_path, sheet_name="Matched Parts", extra_quote_price_keywords=known_quote_price_keywords)
    highlight_prices(output_path, sheet_name="Unmatched Parts", extra_quote_price_keywords=known_quote_price_keywords)

    if progress_label:
        progress_label.config(text=f"\n✅ Done! Prices highlighted and saved to: {output_path}")
        
    else:
        print(f"\n✅ Done! Prices highlighted and saved to: {output_path}")


def launch_gui():
    root = tk.Tk()
    root.title("Spec Comparator Tool")
    root.geometry("600x320")

    frame = ttk.Frame(root, padding=20)
    frame.pack(expand=True, fill="both")

    quote_var = tk.StringVar()
    specs_var = tk.StringVar()
    output_var = tk.StringVar(value="Quote_Spec_Comparison.xlsx")

    def browse_quote():
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if path:
            quote_var.set(path)

    def browse_specs():
        path = filedialog.askdirectory()
        if path:
            specs_var.set(path)

    def browse_output():
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if path:
            output_var.set(path)

    ttk.Label(frame, text="Quote Items Excel File:").grid(row=0, column=0, sticky="w", pady=5)
    ttk.Entry(frame, textvariable=quote_var, width=50).grid(row=0, column=1, padx=5)
    ttk.Button(frame, text="Browse", command=browse_quote).grid(row=0, column=2, padx=5)

    ttk.Label(frame, text="Specs Folder:").grid(row=1, column=0, sticky="w", pady=5)
    ttk.Entry(frame, textvariable=specs_var, width=50).grid(row=1, column=1, padx=5)
    ttk.Button(frame, text="Browse", command=browse_specs).grid(row=1, column=2, padx=5)

    ttk.Label(frame, text="Output Excel File:").grid(row=2, column=0, sticky="w", pady=5)
    ttk.Entry(frame, textvariable=output_var, width=50).grid(row=2, column=1, padx=5)
    ttk.Button(frame, text="Browse", command=browse_output).grid(row=2, column=2, padx=5)

    progress_label = ttk.Label(frame, text="")
    progress_label.grid(row=4, column=1, pady=10)

    def on_run():
        progress_label.config(text="Running...")
        root.update_idletasks()
        try:
            run_comparator(quote_var.get(), specs_var.get(), output_var.get(), progress_label)
        except Exception as e:
            messagebox.showerror("Error", str(e))
            progress_label.config(text="Error occurred.")

    ttk.Button(frame, text="Run Comparison", command=on_run).grid(row=3, column=1, pady=20)

    root.mainloop()


if __name__ == "__main__":
    launch_gui()

def get_diff_chars(a, b):
    # Normalize: lowercase and remove spaces for comparison
    a_raw, b_raw = str(a), str(b)
    a_norm = ''.join(a_raw.lower().split())
    b_norm = ''.join(b_raw.lower().split())
    matcher = difflib.SequenceMatcher(None, a_norm, b_norm)
    diff_a = []
    diff_b = []

    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        # Map normalized indices back to raw string indices
        if tag in ('replace', 'delete'):
            # Find corresponding chars in original a
            raw = ''
            count = 0
            for c in a_raw:
                if c.lower() != ' ':
                    if count >= i1 and count < i2:
                        raw += c
                    count += 1
            diff_a.append(raw)
        if tag in ('replace', 'insert'):
            # Find corresponding chars in original b
            raw = ''
            count = 0
            for c in b_raw:
                if c.lower() != ' ':
                    if count >= j1 and count < j2:
                        raw += c
                    count += 1
            diff_b.append(raw)
    return ''.join(diff_b), ''.join(diff_a)

